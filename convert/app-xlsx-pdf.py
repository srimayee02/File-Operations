import os
from flask import Flask, request, render_template, send_file
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table
from datetime import datetime

app = Flask(__name__)

# Configuration
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['CONVERTED_FOLDER'] = 'converted'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def secure_filename(filename):
    # Remove any unsafe characters from the filename, preserving the dot
    return ''.join(c if c.isalnum() or c == '.' else '_' for c in filename)

def get_unique_filename(filename):
    # Append current date and time to the filename to make it unique
    now = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    return f"{os.path.splitext(filename)[0]}_{now}.pdf"
    

def convert_xlsx_to_pdf_internal(xlsx_path, pdf_path):
    # Load the XLSX file using openpyxl
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # Read data from XLSX and prepare the content for the PDF table
    data = []
    for row in ws.iter_rows():
        row_data = [cell.value for cell in row]
        data.append(row_data)

    # Create a PDF with reportlab
    doc = SimpleDocTemplate(pdf_path, pagesize=letter)
    table = Table(data)
    table.setStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])

    elements = []
    elements.append(table)

    doc.build(elements)

@app.route('/', methods=['GET', 'POST'])
def convert_xlsx_to_pdf():
    if request.method == 'POST':
        # Check if the post request has the file part
        if 'file' not in request.files:
            return "No file part"

        file = request.files['file']

        # If the user does not select a file, the browser might submit an empty part without a filename
        if file.filename == '':
            return "No selected file"

        if file and allowed_file(file.filename):
            # Save the uploaded XLSX file with a unique name
            xlsx_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
            file.save(xlsx_path)

            # Convert XLSX to PDF
            pdf_path = os.path.join(app.config['CONVERTED_FOLDER'], get_unique_filename(file.filename))
            convert_xlsx_to_pdf_internal(xlsx_path, pdf_path)

            return render_template('result-xlsx-pdf.html', pdf_path=pdf_path)

    return render_template('upload-xlsx-pdf.html')

@app.route('/download/<path:pdf_path>')
def download_pdf_xlsx(pdf_path):
    return send_file(pdf_path, as_attachment=True)

if __name__ == "__main__":
    if not os.path.exists(app.config['UPLOAD_FOLDER']):
        os.makedirs(app.config['UPLOAD_FOLDER'])

    if not os.path.exists(app.config['CONVERTED_FOLDER']):
        os.makedirs(app.config['CONVERTED_FOLDER'])

    app.run(debug=True)
